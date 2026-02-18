#!/usr/bin/env python3
"""Convert dadeserasmus.xlsx to data.json with normalized city names."""
import json
import openpyxl

# Normalization map for inconsistent city names
NORMALIZE = {
    "Bilbo": "Bilbao",
    "Grana": "Granada",
    "San Sebastiána": "San Sebastián",
}

wb = openpyxl.load_workbook("dadeserasmus.xlsx")
ws = wb.active

students = []
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    nom, carrera, origen = row
    if nom is None:
        continue
    origen_norm = NORMALIZE.get(origen, origen)
    students.append({
        "nom": nom.strip(),
        "carrera": carrera.strip() if carrera else "",
        "origen": origen_norm.strip() if origen_norm else ""
    })

with open("data.json", "w", encoding="utf-8") as f:
    json.dump(students, f, ensure_ascii=False, indent=2)

print(f"Exported {len(students)} students to data.json")
